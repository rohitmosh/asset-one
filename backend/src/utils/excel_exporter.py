from io import BytesIO
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.models import models

def export_assets_to_excel(instances: list[models.AssetInstance]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Register"

    # Define color fills (using HEX colors matching the guidelines and design)
    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")     # Expired
    orange_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Ending <= 30 Days (using light yellow/orange)
    yellow_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")  # Warning 30-60 Days (light green/yellow)
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers = [
        "Sl No", "IT/OT", "Asset Type", "Asset Group", "Asset", "Asset Identifier", 
        "Asset Description", "Manufacturer", "Model No.", "Serial No.", 
        "Owner", "Contact Details", "Custodian", "User(s)", "Location Name", 
        "Floor Location", "Security Classification", "AMC/Warranty End Date", 
        "Impact on Business Continuity", "Backup Location",
        "Vulnerability of Asset", "Any deviation from company policy"
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 28
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    today = date.today()

    for idx, inst in enumerate(instances, 1):
        asset_type = inst.asset.asset_type.name if inst.asset and inst.asset.asset_type else ""
        asset_group = inst.asset.asset_group.name if inst.asset and inst.asset.asset_group else ""
        asset_name = inst.asset.name if inst.asset else ""
        owner_name = inst.owner.name if inst.owner else ""
        owner_contact = inst.owner.email if inst.owner else ""
        custodian_name = inst.custodian.name if inst.custodian else ""
        assigned_user = inst.assigned_user.name if inst.assigned_user else "N/A"
        
        location_name = f"{inst.location.plant_office} - {inst.location.building}" if inst.location else ""
        floor_location = f"Floor {inst.location.floor or ''}, Room {inst.location.room or ''}" if inst.location else ""
        
        warranty_end = inst.warranty_end_date
        warranty_str = warranty_end.strftime("%Y-%m-%d") if warranty_end else "N/A"
        
        backup_loc = inst.backup_location if inst.backup_available else "No Backup"

        row_data = [
            idx,
            inst.asset.asset_group.domain.upper() if inst.asset and inst.asset.asset_group else "N/A",
            asset_type,
            asset_group,
            asset_name,
            inst.identifier,
            inst.description or "",
            inst.manufacturer or "",
            inst.model_number or "",
            inst.serial_number or "",
            owner_name,
            owner_contact,
            custodian_name,
            assigned_user,
            location_name,
            floor_location,
            inst.security_classification,
            warranty_str,
            inst.business_criticality,
            backup_loc,
            inst.known_vulnerabilities or "None",
            inst.policy_deviations or "None"
        ]
        
        ws.append(row_data)
        row_idx = idx + 1
        ws.row_dimensions[row_idx].height = 20
        
        warranty_cell = ws.cell(row=row_idx, column=18)
        if warranty_end:
            days_remaining = (warranty_end - today).days
            if days_remaining < 0:
                warranty_cell.fill = red_fill
            elif days_remaining <= 30:
                warranty_fill_color = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
                warranty_cell.fill = warranty_fill_color
            elif days_remaining <= 60:
                warranty_fill_color = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                warranty_cell.fill = warranty_fill_color
                
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 2, 6, 17, 18, 19]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    file_stream = BytesIO()
    wb.save(file_stream)
    return file_stream.getvalue()
